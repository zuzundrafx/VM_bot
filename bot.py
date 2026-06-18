import telebot
from telebot import types
import random
import xlrd
import openpyxl
import datetime
import time
import logging
import requests
from urllib.parse import urlencode
import os
from dotenv import load_dotenv
from flask import Flask, request
import sys

load_dotenv()

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

TOKEN = os.getenv('BOT_TOKEN', '7084552505:AAECx4YcUNDJV9SV-Dd4VEddpjyBnR_IBiA')
bot = telebot.TeleBot(TOKEN)

app = Flask(__name__)

hours_row = 0
hours_column = 0

# Глобальные переменные для колонок
spec_column = 0
date_priem_column = 0
stazh_k_column = 0
opit_k_column = 0
compens_column = 0
premia_column = 0
kvart_column = 0
zarplata_pochas_column = 0
za_opit_column = 0
za_stazh_column = 0
zp_column = 0
drivers_column = 0

def download_file_from_yandex():
    """Скачивает файл с Яндекс.Диска и возвращает workbook"""
    try:
        base_url = 'https://cloud-api.yandex.net/v1/disk/public/resources/download?'
        public_key = 'https://disk.yandex.ru/i/cyOG-61LXGmE-g'
        
        final_url = base_url + urlencode(dict(public_key=public_key))
        logger.info(f"Запрос к Яндекс.Диску: {final_url}")
        
        response = requests.get(final_url, timeout=10)
        response.raise_for_status()
        
        download_url = response.json().get('href')
        if not download_url:
            raise Exception("Не удалось получить ссылку для скачивания")
        
        logger.info("Получена ссылка для скачивания")
        
        download_response = requests.get(download_url, timeout=30)
        download_response.raise_for_status()
        
        # Сохраняем файл
        with open('actual_tabel.xlsx', 'wb') as f:
            f.write(download_response.content)
        
        logger.info("Файл успешно скачан")
        
        # Загружаем workbook
        rb1 = openpyxl.load_workbook('actual_tabel.xlsx', data_only=True)
        rb1.iso_dates = True
        return rb1
        
    except requests.exceptions.RequestException as e:
        logger.error(f"Ошибка при скачивании файла: {e}")
        raise
    except Exception as e:
        logger.error(f"Ошибка: {e}")
        raise

def find_user_in_sheet(sheet, user_id):
    """Ищет пользователя в таблице"""
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column, values_only=False):
        for cell in row:
            if cell.value == user_id:
                return cell.row
    return None

@bot.message_handler(commands=['start'])
def handle_message(message):
    cur_version = 1.2
    bot.send_message(message.chat.id, "Текущая версия бота - v" + str(cur_version) + "\n Ваш ID: " + str(message.from_user.id))
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    but1 = types.KeyboardButton("Общая информация")
    but2 = types.KeyboardButton("Табель")
    markup.add(but1, but2)
    bot.reply_to(message, "Здравствуйте, {0.first_name}\n Какая информация вас интересует?".format(message.from_user), parse_mode='html', reply_markup=markup)

@bot.message_handler(func=lambda message: True)
def menu(message):
    if message.chat.type == 'private':
        try:
            if message.text == "Общая информация":
                bot.send_message(message.chat.id, "Ожидайте, собираю информацию")
                
                try:
                    rb1 = download_file_from_yandex()
                    sheet = rb1.active
                    
                    # Поиск колонок
                    for row in sheet.iter_rows(min_row=1, max_row=10, min_col=1, max_col=sheet.max_column, values_only=False):
                        for cell in row:
                            if cell.value == "Специальность":
                                global spec_column
                                spec_column = cell.column
                            elif cell.value == "Дата приема":
                                global date_priem_column
                                date_priem_column = cell.column
                            elif cell.value == "Текущий стаж.коэфф":
                                global stazh_k_column
                                stazh_k_column = cell.column
                            elif cell.value == "Часовая ставка":
                                global opit_k_column
                                opit_k_column = cell.column
                    
                    user_row = find_user_in_sheet(sheet, message.from_user.id)
                    
                    if user_row:
                        bot.send_message(message.chat.id, "Я вас нашел! Собираю данные")
                        bot.send_message(message.chat.id, "ФИО: " + str(sheet.cell(user_row, 2).value))
                        
                        if spec_column and date_priem_column and stazh_k_column and opit_k_column:
                            bot.send_message(message.chat.id, 
                                "Должность: " + str(sheet.cell(user_row, spec_column).value) + 
                                " \nДата трудоустройства: " + str(sheet.cell(user_row, date_priem_column).value) + 
                                "\nКоэффициент за стаж: " + str(sheet.cell(user_row, stazh_k_column).value) + " %" + 
                                "\nЧасовая ставка: " + str(sheet.cell(user_row, opit_k_column).value) + " руб/ч"
                            )
                        else:
                            bot.send_message(message.chat.id, "Не удалось найти все необходимые столбцы в таблице")
                    else:
                        bot.send_message(message.chat.id, "Вас нет в табеле. Обратитесь к администратору.")
                    
                    rb1.close()
                    
                except Exception as e:
                    logger.error(f"Ошибка при обработке: {e}")
                    bot.send_message(message.chat.id, f"Произошла ошибка при загрузке данных. Попробуйте позже.")
                    
            elif message.text == "Табель":
                bot.send_message(message.chat.id, "Ожидайте, готовлю отчет")
                
                try:
                    rb1 = download_file_from_yandex()
                    sheet = rb1.active
                    
                    # Поиск колонок
                    for row in sheet.iter_rows(min_row=1, max_row=10, min_col=1, max_col=sheet.max_column, values_only=False):
                        for cell in row:
                            if cell.value == "Итого часов":
                                global hours_column
                                hours_column = cell.column
                            elif cell.value == "Компенсации":
                                global compens_column
                                compens_column = cell.column
                            elif cell.value == "Премия":
                                global premia_column
                                premia_column = cell.column
                            elif cell.value == "Квартиры":
                                global kvart_column
                                kvart_column = cell.column
                            elif cell.value == "ЗП (почасовка) без учета компенсаций":
                                global zarplata_pochas_column
                                zarplata_pochas_column = cell.column
                            elif cell.value == "ЗП за Опыт":
                                global za_opit_column
                                za_opit_column = cell.column
                            elif cell.value == "ЗП за Стаж":
                                global za_stazh_column
                                za_stazh_column = cell.column
                            elif cell.value == "ЗП (почасовка + премии + стаж + опыт + квартиры)":
                                global zp_column
                                zp_column = cell.column
                            elif cell.value == "Водители":
                                global drivers_column
                                drivers_column = cell.column
                    
                    user_row = find_user_in_sheet(sheet, message.from_user.id)
                    
                    if user_row:
                        bot.send_message(message.chat.id, "Я вас нашел! Формирую отчет за " + str(sheet.cell(1, 2).value))
                        bot.send_message(message.chat.id, "ФИО: " + str(sheet.cell(user_row, 2).value))
                        
                        # Получаем значения с проверкой на None
                        compens = sheet.cell(user_row, compens_column).value if compens_column else 0
                        if compens is None:
                            compens = 0
                        
                        premia = sheet.cell(user_row, premia_column).value if premia_column else 0
                        if premia is None:
                            premia = 0
                        
                        # Формируем отчет
                        message_text = (
                            f"Отработано: {sheet.cell(user_row, hours_column).value if hours_column else 'Н/Д'} часов \n"
                            f"Почасовка: {sheet.cell(user_row, zarplata_pochas_column).value if zarplata_pochas_column else 'Н/Д'} рублей \n"
                            f"Премия: {premia} рублей \n"
                            f"За опыт: {sheet.cell(user_row, za_opit_column).value if za_opit_column else 'Н/Д'} рублей \n"
                            f"За стаж: {sheet.cell(user_row, za_stazh_column).value if za_stazh_column else 'Н/Д'} рублей \n"
                            f"За квартиру: {sheet.cell(user_row, kvart_column).value if kvart_column else 'Н/Д'} рублей \n"
                            f"Водительские: {sheet.cell(user_row, drivers_column).value if drivers_column else 'Н/Д'} рублей \n"
                            f"*Общая ЗП: *{sheet.cell(user_row, zp_column).value if zp_column else 'Н/Д'} рублей \n\n"
                            f"Компенсации: {compens} рублей"
                        )
                        
                        bot.send_message(message.chat.id, message_text, parse_mode='Markdown')
                    else:
                        bot.send_message(message.chat.id, "Вас нет в табеле. Обратитесь к администратору.")
                    
                    rb1.close()
                    
                except Exception as e:
                    logger.error(f"Ошибка при обработке Табеля: {e}")
                    bot.send_message(message.chat.id, f"Произошла ошибка при загрузке данных. Попробуйте позже.")
                    
        except Exception as e:
            logger.error(f"Общая ошибка в menu: {e}")
            bot.send_message(message.chat.id, "Произошла непредвиденная ошибка. Попробуйте позже.")

# Webhook обработчики
@app.route('/')
def index():
    return "Bot is running!"

@app.route('/webhook', methods=['POST'])
def webhook():
    if request.headers.get('content-type') == 'application/json':
        try:
            json_string = request.get_data().decode('utf-8')
            update = telebot.types.Update.de_json(json_string)
            bot.process_new_updates([update])
            return '', 200
        except Exception as e:
            logger.error(f"Ошибка в webhook: {e}")
            return '', 500
    else:
        return 'Invalid content type', 403

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
