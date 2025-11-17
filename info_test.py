import telebot #импорт pyTelegramBotAPI
from telebot import types #также достанем типы
import random #рандом обязательно
import xlrd #библиотка чтения экселевских файлов
import openpyxl #библиотка чтения всех экселевских файлов
import datetime
import time
import logging
import requests
from urllib.parse import urlencode
import datetime
import os
from dotenv import load_dotenv
load_dotenv()

# from telegram import Update
# from telegram.ext import ApplicationBuilder, CommandHandler

# async def start(update: Update, context):
#    await update.message.reply_text('Бот работает!')

#def set_webhook():
#    try:
#        url = f"https://api.telegram.org/bot{TOKEN}/setWebhook"
#        requests.post(url, json={"url": WEBHOOK_URL})
#    except:
#       pass



logger = logging.getLogger(__name__)


from openpyxl import load_workbook, styles

#TOKEN = '7084552505:AAECx4YcUNDJV9SV-Dd4VEddpjyBnR_IBiA'
TOKEN = os.getenv('BOT_TOKEN', '7084552505:AAECx4YcUNDJV9SV-Dd4VEddpjyBnR_IBiA')

bot = telebot.TeleBot(TOKEN)

hours_row = 0
hours_column = 0




@bot.message_handler(commands=['start'])


def handle_message(message):
	cur_version = 1.2 #версия бота
	bot.send_message(message.chat.id, "Текущая версия бота - v" + str(cur_version) + "\n Ваш ID: " + str(message.from_user.id))

# клавиатура
	markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
	but1 = types.KeyboardButton("Общая информация")
	but2 = types.KeyboardButton("Табель")
	markup.add(but1, but2)



	bot.reply_to(message, "Здравствуйте, {0.first_name}\n Какая информация вас интересует?".format(message.from_user)
  ,parse_mode='html',reply_markup=markup)


@bot.message_handler(func=lambda message: True)

def menu(message):

	if message.chat.type == 'private':

		global id_check
		global hours_row
		global hours_column
		global compens_column
		global premia_column
		global zarplata_column
		global spec_column
		global date_priem_column
		global stazh_k_column
		global stazh_column

		if message.text == "Общая информация":

			bot.send_message(message.chat.id, "Ожидайте, собираю информацию")

			base_url = 'https://cloud-api.yandex.net/v1/disk/public/resources/download?'
			public_key = 'https://disk.yandex.ru/i/gFvPIdO1gBanpw'  # Сюда вписываем ссылку на актуальный табель

			#public_key = 'https://disk.yandex.ru/i/iawazDh0f4yqEg' #ссылка на время тестирования

			# Получаем загрузочную ссылку
			final_url = base_url + urlencode(dict(public_key=public_key))
			response = requests.get(final_url)
			download_url = response.json()['href']

			# Загружаем файл и сохраняем его
			download_response = requests.get(download_url)
			with open('actual_tabel.xlsx', 'wb') as f:   # Здесь укажите нужный путь к файлу
			#with open('actual_tabel_testing.xlsx', 'wb') as f:   # Тестирование


				f.write(download_response.content)


			rb1 = openpyxl.load_workbook('actual_tabel.xlsx', data_only=True)

			#rb1 = openpyxl.load_workbook('actual_tabel_testing.xlsx', data_only=True) #Тестирование

			rb1.iso_dates = True

			sheet = rb1.active

			id_check = False

			for row in sheet.iter_rows(min_row=0, min_col=0, max_row=50, max_col=60, values_only=False):

				for cell1 in row:

					if cell1.value == message.from_user.id:
						hours_row = cell1.row
						id_check = True
						bot.send_message(message.chat.id, text="Я вас нашел! Собираю данные")
						bot.send_message(message.chat.id, "ФИО: " + str(sheet.cell(hours_row, 2).value))

					if cell1.value == "Специальность":
						spec_column = cell1.column


					if cell1.value == "Дата приема":
						date_priem_column = cell1.column

						#stazh = sheet.cell(hours_row, date_priem_column).date

					if cell1.value == "Текущий стаж.коэфф":
						stazh_k_column = cell1.column

					if cell1.value == "За опыт!":
						opit_k_column = cell1.column



			if id_check == True:

				bot.send_message(message.chat.id, "Должность: " +  str(sheet.cell(hours_row, spec_column).value) + " \nДата трудоустройства: " + str(sheet.cell(hours_row, date_priem_column).value) + "\nКоэффициент за стаж: " + str(sheet.cell(hours_row, stazh_k_column).value) + " %" + "\nКоэффициент за опыт: " + str(sheet.cell(hours_row, opit_k_column).value) + " %")


		elif message.text == "Табель":



			bot.send_message(message.chat.id, "Ожидайте, готовлю отчет")
			user_id = message.from_user.id
			user_id_str = str(user_id)

			base_url = 'https://cloud-api.yandex.net/v1/disk/public/resources/download?'
			public_key = 'https://disk.yandex.ru/i/gFvPIdO1gBanpw'  # Сюда вписываем ссылку на актуальный табель

			#public_key = 'https://disk.yandex.ru/i/iawazDh0f4yqEg' #ссылка на время тестирования

			# Получаем загрузочную ссылку
			final_url = base_url + urlencode(dict(public_key=public_key))
			response = requests.get(final_url)
			download_url = response.json()['href']

			# Загружаем файл и сохраняем его
			download_response = requests.get(download_url)

			with open('actual_tabel.xlsx', 'wb') as f:   # Здесь укажите нужный путь к файлу

			#with open('actual_tabel_testing.xlsx', 'wb') as f:   # Тестирование

				f.write(download_response.content)


			rb1 = openpyxl.load_workbook('actual_tabel.xlsx', data_only=True) #Стабальная работа бота

			#rb1 = openpyxl.load_workbook('actual_tabel_testing.xlsx', data_only=True) #Тестирование

			rb1.iso_dates = True

			sheet = rb1.active

			id_check = False

			for row in sheet.iter_rows(min_row=0, min_col=0, max_row=50, max_col=60, values_only=False):

				for cell1 in row:

					if cell1.value == message.from_user.id:
						hours_row = cell1.row
						id_check = True
						bot.send_message(message.chat.id, text="Я вас нашел! Формирую отчет за " + sheet.cell(1, 2).value)
						bot.send_message(message.chat.id, "ФИО: " + str(sheet.cell(hours_row, 2).value))

					if cell1.value == "Итого часов":
						hours_column = cell1.column


					if cell1.value == "Компенсации":
						compens_column = cell1.column


					if cell1.value == "Премия":
						premia_column = cell1.column


					if cell1.value == "Квартиры":
						kvart_column = cell1.column


					if cell1.value == "ЗП (почасовка) без учета компенсаций":
						zarplata_pochas_column = cell1.column

					if cell1.value == "ЗП за Опыт":
						za_opit_column = cell1.column

					if cell1.value == "ЗП за Стаж":
						za_stazh_column = cell1.column

					if cell1.value == "ЗП (почасовка + премии + стаж + опыт + квартиры)":
						zp_column = cell1.column

				# Обработка нулевых значений в ячейках таблицы
				if id_check == True:
					if str(sheet.cell(hours_row, compens_column).value) == "None":
						sheet.cell(hours_row, compens_column).value = "0"
					if str(sheet.cell(hours_row, premia_column).value) == "None":
						sheet.cell(hours_row, premia_column).value = "0"

			if id_check == True:

				bot.send_message(message.chat.id, "Отработано: " +  str(sheet.cell(hours_row, hours_column).value) + " часов \nПочасовка: " + str(sheet.cell(hours_row, zarplata_pochas_column).value) + " рублей \nПремия: " + str(sheet.cell(hours_row, premia_column).value) + " рублей \nЗа опыт: " + str(sheet.cell(hours_row, za_opit_column).value) + " рублей \nЗа стаж: " + str(sheet.cell(hours_row, za_stazh_column).value) + " рублей \nЗа квартиру: " + str(sheet.cell(hours_row, kvart_column).value) + " рублей \n*Общая ЗП: *" + str(sheet.cell(hours_row, zp_column).value) + " рублей \n \nКомпенсации: " + str(sheet.cell(hours_row, compens_column).value) + " рублей", parse_mode='Markdown')

				#bot.send_message(message.chat.id, "*жирный текст*, _курсив_ и ~зачеркнутый текст~", parse_mode='Markdown') #Стиль текста в сообщении
				#bot.send_message(message.chat.id, "Премия + стажевые: " + str(sheet.cell(hours_row, premia_column).value) + " рублей")
				#bot.send_message(message.chat.id, "Компенсации: " + str(sheet.cell(hours_row, compens_column).value) + " рублей")
				#bot.send_message(message.chat.id, "Сумма к выплате (без учета компенсаций): " + str(sheet.cell(hours_row, zarplata_column).value) + " рублей")

				rb1.close()

#дублирование сообщений пользователя
#@bot.message_handler(content_types=['text'])
#def handle_message(message):
#	bot.reply_to(message, message.text)


#Старый метод бесперебойной работы бота
#bot.polling(none_stop=True)

#альтернативный метод обхода падения бота
if __name__ == '__main__':
	bot.infinity_polling(none_stop=True)

#if __name__ == '__main__':
    # Для Railway используем порт из переменных окружения
#    port = int(os.environ.get('PORT', 5000))
#    app.run(host='0.0.0.0', port=port)